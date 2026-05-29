"""
07_export.py 
Generates the 14-sheet Excel workbook
Sheets: 00_TOC | 01_Raw | 02_FL21 | 03_FL22 | 04_Scale | 05_Variables |
        06_Correlations | 07_Compliance | 08_Dropout | 09_Tone_Comparisons |
        10_Tone_Effects | 11_Variable_Relationships | 12_Mediation_Moderation |
        13_Quality_Progression | 14_Word_Frequencies
"""
import sys, json
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import numpy as np
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from config import (OUTPUT_FILE, OUTPUT_DIR, COLOR_FL21, COLOR_FL22,
                    COLOR_HDR, COLOR_ALT, SCORE_COLORS, ALL_Q_COLS, NUM_COLS,
                    EXCEL_LETTERS, Q_LABELS, CONSTRUCTS, COL_SCALE, LIKERT_7, CAPABLE_7)

# ── Style constants ────────────────────────────────────────────
THIN = Border(left=Side(style="thin",color="BFBFBF"),right=Side(style="thin",color="BFBFBF"),
               top=Side(style="thin",color="BFBFBF"),bottom=Side(style="thin",color="BFBFBF"))
C_EVAL="2E75B6"; C_QUALITY="1A6634"; C_METRICS="7B3F00"
C_ACTION="6C3483"; C_PSY="C0392B"; C_CONV="1A5276"

# ── Base helpers ───────────────────────────────────────────────
def hdr(c,bg=COLOR_HDR,fc="FFFFFF",sz=10,bold=True):
    c.font=Font(bold=bold,color=fc,name="Arial",size=sz)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=THIN

def dat(c,val,center=True,alt=False,bold=False,fc="000000",fmt=None):
    v=None if (isinstance(val,float) and np.isnan(val)) else val
    if isinstance(v,bool): v="Yes" if v else "No"
    c.value=v; c.font=Font(name="Arial",size=9,bold=bold,color=fc)
    c.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=True)
    c.border=THIN
    if alt: c.fill=PatternFill("solid",start_color=COLOR_ALT)
    if fmt: c.number_format=fmt

def pval_cell(c,p):
    c.border=THIN; c.alignment=Alignment(horizontal="center",vertical="center")
    if p is None or (isinstance(p,float) and np.isnan(p)):
        c.value="n/a"; c.font=Font(name="Arial",size=9); return
    if   p<0.001: c.value=f"{p:.4f}***"; c.fill=PatternFill("solid",start_color="1A6634"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif p<0.01:  c.value=f"{p:.4f}**";  c.fill=PatternFill("solid",start_color="52BE80"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif p<0.05:  c.value=f"{p:.4f}*";   c.fill=PatternFill("solid",start_color="F9E79F"); c.font=Font(name="Arial",size=9,bold=True,color="7D6608")
    else:         c.value=f"{p:.4f} ns"; c.fill=PatternFill("solid",start_color="FADBD8"); c.font=Font(name="Arial",size=9,color="922B21")

def delta_cell(c,d):
    c.border=THIN; c.alignment=Alignment(horizontal="center",vertical="center")
    if d is None or (isinstance(d,float) and np.isnan(d)): c.value="n/a"; return
    c.value=round(d,3); c.number_format="0.000"
    if   d>0.05:  c.fill=PatternFill("solid",start_color="D5F5E3"); c.font=Font(name="Arial",size=9,bold=True,color="1A6634")
    elif d<-0.05: c.fill=PatternFill("solid",start_color="FADBD8"); c.font=Font(name="Arial",size=9,bold=True,color="922B21")
    else:         c.fill=PatternFill("solid",start_color="F2F3F4"); c.font=Font(name="Arial",size=9,color="555555")

def section_title(ws,text,bg,ncols,row,sz=11):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text)
    c.font=Font(bold=True,color="FFFFFF",name="Arial",size=sz)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[row].height=22

def summary_box(ws,text,row,ncols,bg="EBF5FB"):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text)
    c.font=Font(italic=True,name="Arial",size=9,color="1F4E79")
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    c.border=THIN; ws.row_dimensions[row].height=30

def write_ttest_table(ws,rows,start_row,bg_hdr):
    hdrs=["Variable","Label","N FL_21","Mean FL_21","SD FL_21",
          "N FL_22","Mean FL_22","SD FL_22","Δ (21−22)","t-stat","p-value","Cohen's d","Effect size"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(start_row,j,value=h),bg=bg_hdr,sz=9)
    ws.row_dimensions[start_row].height=24
    for i,r in enumerate(rows):
        rr=start_row+1+i; alt=(i%2==0)
        dat(ws.cell(rr,1),r.get("variable",""),center=False,alt=alt,bold=True)
        dat(ws.cell(rr,2),r.get("label",""),center=False,alt=alt)
        dat(ws.cell(rr,3),r.get("n_fl21"),alt=alt)
        dat(ws.cell(rr,4),r.get("mean_fl21"),alt=alt,fmt="0.000")
        dat(ws.cell(rr,5),r.get("sd_fl21"),alt=alt,fmt="0.000")
        dat(ws.cell(rr,6),r.get("n_fl22"),alt=alt)
        dat(ws.cell(rr,7),r.get("mean_fl22"),alt=alt,fmt="0.000")
        dat(ws.cell(rr,8),r.get("sd_fl22"),alt=alt,fmt="0.000")
        delta_cell(ws.cell(rr,9),r.get("delta"))
        dat(ws.cell(rr,10),r.get("t"),alt=alt,fmt="0.0000")
        pval_cell(ws.cell(rr,11),r.get("p"))
        dat(ws.cell(rr,12),r.get("cohens_d"),alt=alt,fmt="0.000")
        dat(ws.cell(rr,13),r.get("effect_size",""),alt=alt)
    widths=[22,38,7,11,9,7,11,9,12,10,14,11,12]
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return start_row+1+len(rows)

def write_chi2_table(ws,rows,start_row,bg_hdr):
    hdrs=["Variable","Label","% FL_21","% FL_22","% Total","Δ% (21−22)","Chi²","p-value"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(start_row,j,value=h),bg=bg_hdr,sz=9)
    for i,r in enumerate(rows):
        rr=start_row+1+i; alt=(i%2==0)
        dat(ws.cell(rr,1),r.get("variable",""),center=False,alt=alt,bold=True)
        dat(ws.cell(rr,2),r.get("label",""),center=False,alt=alt)
        dat(ws.cell(rr,3),r.get("pct_fl21"),alt=alt,fmt="0.0")
        dat(ws.cell(rr,4),r.get("pct_fl22"),alt=alt,fmt="0.0")
        dat(ws.cell(rr,5),r.get("pct_total"),alt=alt,fmt="0.0")
        delta_cell(ws.cell(rr,6),r.get("delta_pct"))
        dat(ws.cell(rr,7),r.get("chi2"),alt=alt,fmt="0.0000")
        pval_cell(ws.cell(rr,8),r.get("p"))
    return start_row+1+len(rows)

def insert_png(ws,path,anchor,w_cm=18,h_cm=9):
    if not Path(path).exists(): return
    img=XLImage(str(path)); img.width=int(w_cm*37.795); img.height=int(h_cm*37.795)
    ws.add_image(img,anchor)

def fmt_coef(coef,sig):
    """Format coefficient with significance stars in one cell."""
    if coef is None: return "—"
    stars={"***":"***","**":"**","*":"*"}.get(sig,"")
    return f"{coef:.4f}{stars}" if coef is not None else "—"

def pval_sig_cell(c,p,sig):
    c.border=THIN; c.alignment=Alignment(horizontal="center",vertical="center")
    if p is None: c.value="—"; c.font=Font(name="Arial",size=9); return
    txt=f"{p:.4f}{sig}" if sig not in ("ns","n/a") else f"{p:.4f}"
    c.value=txt
    if   sig=="***": c.fill=PatternFill("solid",start_color="1A6634"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif sig=="**":  c.fill=PatternFill("solid",start_color="52BE80"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif sig=="*":   c.fill=PatternFill("solid",start_color="F9E79F"); c.font=Font(name="Arial",size=9,bold=True,color="7D6608")
    else:            c.fill=PatternFill("solid",start_color="FADBD8"); c.font=Font(name="Arial",size=9,color="922B21")

# ── Grand mediation summary table (one row per model) ─────────
def write_grand_mediation_table(ws,models,start_row,bg_hdr,show_bloc=True):
    """
    models: list of dicts with keys:
      label, bloc (optional), iv_label, mediator_label, dv_label,
      n, path_a, path_b, path_c, path_cp, indirect, mediation_type
    path_x: dict with coef, p, sig
    indirect: dict with coef, ci_low, ci_up, significant
    """
    cols=[]
    if show_bloc: cols.append("Bloc")
    cols+=["IV","Mediator","DV","n",
           "a (IV→M)","b (M→DV|IV)","c (total)","c' (direct)",
           "Indirect","CI low","CI high","Sig.","Type"]
    for j,h in enumerate(cols,1): hdr(ws.cell(start_row,j,value=h),bg=bg_hdr,sz=9)
    ws.row_dimensions[start_row].height=28
    row=start_row+1
    prev_bloc=""
    for i,m in enumerate(models):
        alt=(i%2==0)
        col_offset=0
        if show_bloc:
            bloc=m.get("bloc","")
            if bloc!=prev_bloc:
                dat(ws.cell(row,1),bloc,center=False,alt=alt,bold=True,fc="1F4E79")
                ws.cell(row,1).fill=PatternFill("solid",start_color="D6EAF8")
            else:
                dat(ws.cell(row,1),"",alt=alt)
            prev_bloc=bloc; col_offset=1
        dat(ws.cell(row,1+col_offset),m.get("iv_label",""),center=False,alt=alt)
        dat(ws.cell(row,2+col_offset),m.get("mediator_label",""),center=False,alt=alt)
        dat(ws.cell(row,3+col_offset),m.get("dv_label",""),center=False,alt=alt)
        dat(ws.cell(row,4+col_offset),m.get("n"),alt=alt)
        # Paths: coef + sig in same cell
        for k,path_key in enumerate(["path_a","path_b","path_c","path_cp"]):
            p_d=m.get(path_key,{})
            c_cell=ws.cell(row,5+col_offset+k)
            val=fmt_coef(p_d.get("coef"),p_d.get("sig","ns"))
            pval_sig_cell(c_cell,p_d.get("p"),p_d.get("sig","ns"))
            c_cell.value=val if val!="—" else c_cell.value
        # Indirect
        ind=m.get("indirect",{})
        sig_ind=ind.get("significant",False)
        dat(ws.cell(row,9+col_offset),ind.get("coef"),alt=alt,fmt="0.0000")
        # CI
        ci_l=ws.cell(row,10+col_offset,value=ind.get("ci_low"))
        ci_l.number_format="0.0000"; ci_l.border=THIN
        ci_l.alignment=Alignment(horizontal="center",vertical="center")
        ci_l.fill=PatternFill("solid",start_color="D5F5E3" if sig_ind else "FADBD8")
        ci_l.font=Font(name="Arial",size=9,color="1A6634" if sig_ind else "922B21")
        ci_u=ws.cell(row,11+col_offset,value=ind.get("ci_up"))
        ci_u.number_format="0.0000"; ci_u.border=THIN
        ci_u.alignment=Alignment(horizontal="center",vertical="center")
        ci_u.fill=PatternFill("solid",start_color="D5F5E3" if sig_ind else "FADBD8")
        ci_u.font=Font(name="Arial",size=9,color="1A6634" if sig_ind else "922B21")
        sc=ws.cell(row,12+col_offset,value="✓ SIG" if sig_ind else "✗ ns")
        sc.border=THIN; sc.alignment=Alignment(horizontal="center",vertical="center")
        sc.fill=PatternFill("solid",start_color="1A6634" if sig_ind else "FADBD8")
        sc.font=Font(name="Arial",size=9,bold=sig_ind,color="FFFFFF" if sig_ind else "922B21")
        dat(ws.cell(row,13+col_offset),m.get("mediation_type",""),center=False,alt=alt)
        ws.row_dimensions[row].height=18; row+=1
    # column widths
    w_start=[20] if show_bloc else []
    widths=w_start+[28,22,28,5,13,13,13,13,10,10,10,10,22]
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return row

def med_to_row(key,item,bloc=None):
    """Convert a mediation result dict to a flat row for grand table."""
    m=item.get("model",{}) if "model" in item else item
    if not m or "error" in m:
        return {"bloc":bloc or "","label":item.get("label",""),
                "iv_label":m.get("iv","") if m else "","mediator_label":"","dv_label":"",
                "n":m.get("n","ERROR") if m else "ERROR",
                "path_a":{},"path_b":{},"path_c":{},"path_cp":{},
                "indirect":{},"mediation_type":"Error / insufficient data"}
    return {"bloc":bloc or item.get("bloc",""),
            "label":item.get("label",""),
            "iv_label":m.get("iv",""),"mediator_label":m.get("mediator",""),"dv_label":m.get("dv",""),
            "n":m.get("n"),
            "path_a":m.get("path_a",{}),"path_b":m.get("path_b",{}),
            "path_c":m.get("path_c",{}),"path_cp":m.get("path_cp",{}),
            "indirect":m.get("indirect",{}),"mediation_type":m.get("mediation_type","")}

def write_grand_regression_table(ws,reg_items,start_row,bg_hdr):
    """
    reg_items: list of (group_label, group_color, [result_dicts])
    Each result: {"label":..., "model":{"n":..,"r2":..,"predictors":[{"variable","label","b","se","t","p","sig"}]}}
    One row per predictor; n and R² shown only on first predictor of each model.
    """
    hdrs=["Group","Model / DV","IV (predictor)","n","R²","b","SE","t","p-value"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(start_row,j,value=h),bg=bg_hdr,sz=9)
    ws.row_dimensions[start_row].height=28
    row=start_row+1
    prev_group=""
    for group_lbl,group_color,items in reg_items:
        for item in items:
            model=item.get("model",{})
            if not model: continue
            preds=model.get("predictors",[])
            if not preds: continue
            n_val=model.get("n"); r2_val=model.get("r2"); model_lbl=item.get("label","")
            for k,pred in enumerate(preds):
                alt=(row%2==0)
                # Group
                if group_lbl!=prev_group:
                    g_cell=ws.cell(row,1,value=group_lbl)
                    g_cell.font=Font(bold=True,color="FFFFFF",name="Arial",size=9)
                    g_cell.fill=PatternFill("solid",start_color=group_color)
                    g_cell.border=THIN; g_cell.alignment=Alignment(horizontal="left",vertical="center")
                    prev_group=group_lbl
                else: dat(ws.cell(row,1),"",alt=alt)
                dat(ws.cell(row,2),model_lbl if k==0 else "",center=False,alt=alt,bold=(k==0))
                dat(ws.cell(row,3),pred.get("label",""),center=False,alt=alt)
                dat(ws.cell(row,4),n_val if k==0 else "",alt=alt)
                dat(ws.cell(row,5),r2_val if k==0 else "",alt=alt,fmt="0.000")
                dat(ws.cell(row,6),pred.get("b"),alt=alt,fmt="0.0000")
                dat(ws.cell(row,7),pred.get("se"),alt=alt,fmt="0.0000")
                dat(ws.cell(row,8),pred.get("t"),alt=alt,fmt="0.0000")
                pval_sig_cell(ws.cell(row,9),pred.get("p"),pred.get("sig","ns"))
                ws.row_dimensions[row].height=16; row+=1
    widths=[22,40,30,6,8,10,10,10,14]
    for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
    return row

# ================================================================
# LOAD ALL DATA
# ================================================================
def load_all():
    def jload(name):
        p=OUTPUT_DIR/name
        if p.exists():
            with open(p,encoding="utf-8") as f: return json.load(f)
        return {}
    return {
        "df_clean":   pd.read_json(OUTPUT_DIR/"df_clean.json")   if (OUTPUT_DIR/"df_clean.json").exists()  else pd.DataFrame(),
        "df_fl21":    pd.read_json(OUTPUT_DIR/"df_fl21.json")    if (OUTPUT_DIR/"df_fl21.json").exists()   else pd.DataFrame(),
        "df_fl22":    pd.read_json(OUTPUT_DIR/"df_fl22.json")    if (OUTPUT_DIR/"df_fl22.json").exists()   else pd.DataFrame(),
        "df_coding":  pd.read_json(OUTPUT_DIR/"df_coding.json")  if (OUTPUT_DIR/"df_coding.json").exists() else pd.DataFrame(),
        "df_ai":      pd.read_json(OUTPUT_DIR/"df_ai.json")      if (OUTPUT_DIR/"df_ai.json").exists()     else pd.DataFrame(),
        "df_agg":     pd.read_json(OUTPUT_DIR/"df_agg.json")     if (OUTPUT_DIR/"df_agg.json").exists()    else pd.DataFrame(),
        "df_prog":    pd.read_json(OUTPUT_DIR/"df_progression.json") if (OUTPUT_DIR/"df_progression.json").exists() else pd.DataFrame(),
        "correlations":    jload("correlations_full.json"),
        "tone_comparisons":jload("tone_comparisons.json"),
        "mediation_tone":  jload("mediation_tone.json"),
        "regression_noton":jload("regression_noton.json"),
        "mediation_noton": jload("mediation_noton.json"),
        "dropout":         jload("dropout_corrected.json"),
        "wordfreq":        jload("wordfreq.json"),
        "synthesis":       jload("synthesis.json"),
    }

# ================================================================
# SHEET 00 — TABLE OF CONTENTS
# ================================================================
def build_sheet0(wb):
    ws=wb.active; ws.title="00_Table_of_Contents"
    section_title(ws,"Master Thesis — Full Analysis | Table of Contents",COLOR_HDR,4,1)
    toc=[
        ("01","Raw_Data_Cleaned","Raw Qualtrics data after filters (Progress=100, n_bot_msgs≥1)","—"),
        ("02","FL21_Friendly","All responses — Friendly tone condition (FL_21)","—"),
        ("03","FL22_Professional","All responses — Professional tone condition (FL_22)","—"),
        ("04","Scale_Coding","Likert / Capable / Amount scale definitions and coding","—"),
        ("05","Variable_Dictionary","Definition, source, scale, and scoring of every retained variable","—"),
        ("06","Correlations","Pearson correlation matrix (significant pairs + bloc zooms by construct)","Exploration"),
        ("07","Chatbot_Compliance","Chatbot tone compliance verification — markers, scores, t-test","Manipulation check"),
        ("08","Dropout_Analysis","End type distribution, dropout by tone, E1/E2 by end type","Q1.5, Q5.1"),
        ("09","Tone_Comparisons","All DVs compared by tone: t-tests (continuous) + chi² (binary)","Q1.1–Q1.5"),
        ("10","Tone_Effects","Mediations and regressions where Tone = IV","Q2.4, Q2.5, Q3.4, Q4.5 + emotion mediations"),
        ("11","Variable_Relationships","Regressions between variables (Tone not IV)","Q2.1–Q2.3, Q3.1–Q3.2, Q4.1, Q5.2–Q5.3"),
        ("12","Mediation_Moderation","Mediations between variables (Tone not IV)","Q2.6, Q3.3, Q4.4 + emotion mediations"),
        ("13","Quality_Progression","Quality score per conversation turn — FL_21 vs FL_22 curve","Longitudinal quality"),
        ("14","Word_Frequencies","Top words + TF-IDF distinctive terms + word clouds","Qualitative illustration"),
    ]
    hdrs=["Sheet","Name","Content","Research questions answered"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(2,j,value=h),bg="2C3E50",sz=10)
    for i,(num,name,content,rq) in enumerate(toc):
        r=3+i; alt=(i%2==0)
        dat(ws.cell(r,1),num,alt=alt,bold=True)
        dat(ws.cell(r,2),name,center=False,alt=alt,bold=True)
        dat(ws.cell(r,3),content,center=False,alt=alt)
        dat(ws.cell(r,4),rq,center=False,alt=alt)
        ws.row_dimensions[r].height=22
    for j,w in enumerate([8,28,70,45],1): ws.column_dimensions[get_column_letter(j)].width=w

# ================================================================
# SHEETS 01–04 (unchanged logic, abbreviated)
# ================================================================
def build_raw_sheet(wb, df, title, sheet_name, bg):
    ws=wb.create_sheet(sheet_name)
    if df.empty: return
    cols=[c for c in df.columns if not c.endswith("_num")]
    section_title(ws,f"{title} (n={len(df)})",bg,len(cols),1)
    for j,col in enumerate(cols,1): hdr(ws.cell(2,j,value=col),bg="2C3E50",sz=8)
    for i,(_,row) in enumerate(df[cols].iterrows()):
        for j,v in enumerate(row,1): dat(ws.cell(3+i,j),v,center=(j>2),alt=(i%2==0))
    ws.freeze_panes="A3"

def build_sheet4(wb, df_coding):
    ws=wb.create_sheet("04_Scale_Coding")
    section_title(ws,"Likert scale coding reference",COLOR_HDR,10,1)
    defs=[("Likert 1-7 (agreement)","2E75B6",["Strongly\ndisagree","Disagree","Somewhat\ndisagree","Neither","Somewhat\nagree","Agree","Strongly\nagree"],"evaluation_1-6, Perceived Manipulati_1-4"),
          ("Capable 1-7","1A6634",["Not at all\ncapable","Very slightly\ncapable","Slightly\ncapable","Neither","Somewhat\ncapable","Capable","Very\ncapable"],"Competence_1-2"),
          ("Amount 1-7","7B3F00",["Not\nat all","Very\nlittle","A\nlittle","Moderate\namount","Quite\na bit","A\nlot","A great\ndeal"],"Moral Responsibility_1-4, Sense of Independenc_1-2")]
    row=3
    for sname,color,labels,cols_info in defs:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=9)
        c=ws.cell(row,1,value=sname); c.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
        c.fill=PatternFill("solid",start_color=color); c.alignment=Alignment(horizontal="left",vertical="center")
        ws.row_dimensions[row].height=22; row+=1
        ws.cell(row,1,value=f"Columns: {cols_info}").font=Font(italic=True,name="Arial",size=9,color="555555"); row+=1
        ws.cell(row,1,value="Score →").font=Font(bold=True,name="Arial",size=9)
        for j,(score,label) in enumerate(zip(range(1,8),labels),2):
            bg,fc=SCORE_COLORS[score]
            c1=ws.cell(row,j,value=score); c1.font=Font(bold=True,color=fc,name="Arial",size=11)
            c1.fill=PatternFill("solid",start_color=bg); c1.alignment=Alignment(horizontal="center",vertical="center"); c1.border=THIN
            c2=ws.cell(row+1,j,value=label); c2.font=Font(name="Arial",size=8)
            c2.fill=PatternFill("solid",start_color="F8F9FA")
            c2.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c2.border=THIN
            ws.column_dimensions[get_column_letter(j)].width=14
        ws.row_dimensions[row].height=20; ws.row_dimensions[row+1].height=34; row+=3

# ================================================================
# SHEET 05 — VARIABLE DICTIONARY
# ================================================================
def build_sheet5(wb):
    ws=wb.create_sheet("05_Variable_Dictionary")
    section_title(ws,"Variable Dictionary — all retained variables",COLOR_HDR,6,1)
    variables=[
        ("EVALUATION QUESTIONNAIRE (Likert 1-7)",C_EVAL,None,None,None,None),
        ("evaluation_1_num","E1 Required effort","Likert 1-7","Questionnaire","Continuous","1=no effort, 7=very high effort. Higher = more effort perceived. REVERSED item."),
        ("evaluation_2_num","E2 Engagement felt","Likert 1-7","Questionnaire","Continuous","1=strongly disagree, 7=strongly agree. 'I felt engaged while answering this survey.'"),
        ("evaluation_3_num","E3 Chatbot appreciation","Likert 1-7","Questionnaire","Continuous","'I enjoyed interacting with the chatbot.'"),
        ("evaluation_4_num","E4 Conversation utility","Likert 1-7","Questionnaire","Continuous","'I found the conversation useful for providing my feedback.'"),
        ("evaluation_5_num","E5 Reuse intention","Likert 1-7","Questionnaire","Continuous","'I would consider using this chatbot again.'"),
        ("evaluation_6_num","E6 Chatbot preference","Likert 1-7","Questionnaire","Continuous","'I would prefer to provide feedback to a chatbot rather than a human.'"),
        ("PERCEIVED MANIPULATION (Likert 1-7)",C_PSY,None,None,None,None),
        ("Perceived Manipulati_1_num","PM1 Freedom threat","Likert 1-7","Questionnaire","Continuous","'The chatbot threatened my freedom to choose.' Based on psychological reactance theory."),
        ("Perceived Manipulati_2_num","PM2 Decision override","Likert 1-7","Questionnaire","Continuous","'The chatbot tried to make a decision for me.'"),
        ("Perceived Manipulati_3_num","PM3 Manipulation","Likert 1-7","Questionnaire","Continuous","'The chatbot tried to manipulate me.'"),
        ("Perceived Manipulati_4_num","PM4 Pressure","Likert 1-7","Questionnaire","Continuous","'The chatbot tried to pressure me.'"),
        ("AI COMPETENCE (Capable scale 1-7)",C_PSY,None,None,None,None),
        ("Competence_1_num","Comp1 Skills judgment","Capable 1-7","Questionnaire","Continuous","'How capable is an AI of judging someone's abilities?' 1=not at all, 7=very capable."),
        ("Competence_2_num","Comp2 Moral judgment","Capable 1-7","Questionnaire","Continuous","'How capable is an AI of judging how moral someone's behavior is?'"),
        ("MORAL RESPONSIBILITY (Amount 1-7)",C_PSY,None,None,None,None),
        ("Moral Responsibility_1_num","MR1 AI harm wrong","Amount 1-7","Questionnaire","Continuous","'How morally wrong would it be for this AI to harm a person?'"),
        ("Moral Responsibility_2_num","MR2 AI responsible","Amount 1-7","Questionnaire","Continuous","'To what extent would this AI deserve to be held morally responsible for a negative outcome?'"),
        ("Moral Responsibility_3_num","MR3 Human harm AI","Amount 1-7","Questionnaire","Continuous","'How morally wrong would it be for someone to harm this AI?'"),
        ("Moral Responsibility_4_num","MR4 AI moral concern","Amount 1-7","Questionnaire","Continuous","'To what extent does this AI deserve to be treated with moral concern?'"),
        ("AI SENSE OF INDEPENDENCE (Amount 1-7)",C_PSY,None,None,None,None),
        ("Sense of Independenc_1_num","Ind1 AI plans & goals","Amount 1-7","Questionnaire","Continuous","'To what extent can this AI make plans and work towards goals?'"),
        ("Sense of Independenc_2_num","Ind2 AI self-control","Amount 1-7","Questionnaire","Continuous","'To what extent can this AI exercise self-control?'"),
        ("AI QUALITY SCORES (GPT-4o-mini rated, 1-5)",C_QUALITY,None,None,None,None),
        ("quality_global","Quality global","1-5","AI (GPT-4o-mini)","Continuous","1=1-3 word answers no development, 3=adequate incomplete, 5=rich nuanced multiple examples"),
        ("quality_precision","Quality precision","1-5","AI (GPT-4o-mini)","Continuous","1=very vague, 5=every statement grounded in a fact or concrete example"),
        ("quality_examples","Quality examples","1-5","AI (GPT-4o-mini)","Continuous","1=no concrete example, 5=multiple detailed examples illustrating different aspects"),
        ("quality_relevance","Quality relevance","1-5","AI (GPT-4o-mini)","Continuous","1=answers don't address questions, 5=answers perfectly, anticipates follow-up"),
        ("quality_richness","Quality richness","1-5","AI (GPT-4o-mini)","Continuous","1=limited vocabulary repetitive sentences, 5=rich vocabulary precise register fine nuances"),
        ("ACTIONABILITY & CONTENT (Binary, AI-rated)",C_ACTION,None,None,None,None),
        ("action_concrete_pb","Concrete problem","0/1","AI (GPT-4o-mini)","Binary","1=participant describes a precise problem with context (when X happens Y occurs because Z)"),
        ("action_advice","Applicable advice","0/1","AI (GPT-4o-mini)","Binary","1=participant formulates an improvement a product team could implement directly"),
        ("action_use_case","Precise use case","0/1","AI (GPT-4o-mini)","Binary","1=participant describes precise usage context (when/where/how/why)"),
        ("content_emotion","Emotion expressed","0/1","AI (GPT-4o-mini)","Binary","1=participant expresses a clear emotion (frustration, satisfaction) with at least some context"),
        ("TEXT METRICS (NLTK)",C_METRICS,None,None,None,None),
        ("avg_words_per_msg","Avg words/message","Count","NLTK","Continuous","Average alphabetic tokens per message. Excludes punctuation."),
        ("avg_word_len","Avg word length","Characters","NLTK","Continuous","Average characters per word. Proxy for lexical complexity."),
        ("n_turns","Number of turns","Count","Auto","Continuous","Number of participant message turns in the conversation."),
        ("CONVERSATION",C_CONV,None,None,None,None),
        ("end_type","End type","Categorical","Auto+AI","Categorical","proper_end=closing phrase detected; early_dropout=≤3 turns AND ≤5 words; minimal_response=last msg ≤3 words"),
        ("breakpoint_exists","Breakpoint detected","0/1","AI (GPT-4o-mini)","Binary","1=at least 2 consecutive turns show lasting quality degradation"),
        ("breakpoint_turn","Breakpoint turn","Turn #","AI (GPT-4o-mini)","Continuous","Turn where quality started consistently declining. Only meaningful when breakpoint_exists=1."),
        ("CHATBOT COMPLIANCE (AI-rated)",C_CONV,None,None,None,None),
        ("bot_score_friendly","Bot friendly score","1-5","AI (GPT-4o-mini)","Continuous","1=cold/distant, 3=neutral, 5=very warm/enthusiastic with emojis and informal address"),
        ("bot_score_professional","Bot professional score","1-5","AI (GPT-4o-mini)","Continuous","1=very casual, 3=neutral, 5=very formal systematic formal address no emojis"),
        ("bot_compliance_score","Bot compliance score","1-5","AI (GPT-4o-mini)","Continuous","1=completely opposite to brief, 3=partially compliant, 5=perfectly compliant"),
        ("bot_coherence_score","Bot tone coherence","1-5","AI (GPT-4o-mini)","Continuous","Consistency of tone throughout conversation. 1=major drifts, 5=perfectly consistent"),
    ]
    hdrs=["Variable","Short label","Scale","Source","Type","Definition & scoring"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(2,j,value=h),bg="2C3E50",sz=10)
    row=3
    for item in variables:
        if item[2] is None:
            section_title(ws,item[0],item[1],6,row); row+=1
        else:
            alt=((row-3)%2==0)
            for j,v in enumerate(item,1): dat(ws.cell(row,j),v,center=(j>2 and j<6),alt=alt,bold=(j==1))
            ws.row_dimensions[row].height=35; row+=1
    for j,w in enumerate([28,22,14,18,12,70],1): ws.column_dimensions[get_column_letter(j)].width=w

# ================================================================
# SHEET 06 — CORRELATIONS (sig pairs + bloc zooms only)
# ================================================================
def build_sheet6(wb, corr_data):
    ws=wb.create_sheet("06_Correlations")
    section_title(ws,"Correlations — Significant pairs + bloc zooms",COLOR_HDR,9,1)
    summary_box(ws,"Pearson r. |r|<0.3=weak, 0.3–0.5=moderate, >0.5=strong. Green=positive, Red=negative. Significance: * p<.05  ** p<.01  *** p<.001",2,9)
    if not corr_data: ws.cell(3,1,value="Run 08_new_analyses.py first"); return

    sig_pairs=corr_data.get("significant_pairs",[])
    row=4
    section_title(ws,f"All significant pairs sorted by |r| (n={len(sig_pairs)} pairs with p<.05)","2C3E50",9,row); row+=1
    for j,h in enumerate(["Variable X","Label X","Variable Y","Label Y","r","p","Sig.","Strength","Direction"],1):
        hdr(ws.cell(row,j,value=h),bg="2C3E50",sz=9)
    row+=1
    for i,pair in enumerate(sig_pairs[:60]):
        alt=(i%2==0)
        dat(ws.cell(row,1),pair["var_x"],center=False,alt=alt,bold=True)
        dat(ws.cell(row,2),pair["label_x"],center=False,alt=alt)
        dat(ws.cell(row,3),pair["var_y"],center=False,alt=alt,bold=True)
        dat(ws.cell(row,4),pair["label_y"],center=False,alt=alt)
        rv=pair["r"]
        rc=ws.cell(row,5,value=rv); rc.number_format="0.000"; rc.border=THIN
        rc.alignment=Alignment(horizontal="center",vertical="center")
        if abs(rv)>=0.5: rc.fill=PatternFill("solid",start_color="1A6634"); rc.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
        elif abs(rv)>=0.3: rc.fill=PatternFill("solid",start_color="A9DFBF"); rc.font=Font(name="Arial",size=9)
        elif rv<0: rc.fill=PatternFill("solid",start_color="F1948A"); rc.font=Font(name="Arial",size=9)
        pval_cell(ws.cell(row,6),pair["p"])
        dat(ws.cell(row,7),pair["sig"],alt=alt,bold=True)
        dat(ws.cell(row,8),pair["strength"],alt=alt)
        dat(ws.cell(row,9),pair["direction"],alt=alt)
        row+=1

    # Bloc zooms
    for bname,bdata in corr_data.get("bloc_matrices",{}).items():
        row+=2; blabels=bdata.get("labels",[]); bd=bdata.get("data",[])
        section_title(ws,f"Zoom: {bname}","2C3E50",len(blabels)+1,row); row+=1
        ws.cell(row,1,value="↓/→").font=Font(bold=True,name="Arial",size=8)
        for j,lbl in enumerate(blabels,2):
            c=ws.cell(row,j,value=lbl); c.font=Font(bold=True,name="Arial",size=8); c.border=THIN
            c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.fill=PatternFill("solid",start_color="EBF5FB")
            ws.column_dimensions[get_column_letter(j)].width=max(8,len(lbl)//2)
        row+=1
        for rd in bd:
            ws.cell(row,1,value=rd.get("label","")).font=Font(name="Arial",size=8,bold=True); ws.cell(row,1).border=THIN
            for j,cell in enumerate(rd.get("cells",[]),2):
                c=ws.cell(row,j); c.border=THIN; c.alignment=Alignment(horizontal="center",vertical="center")
                rv=cell.get("r"); pv=cell.get("p")
                if rv is None: c.value="—"; c.font=Font(name="Arial",size=8); continue
                if rv==1.0: c.value="1"; c.fill=PatternFill("solid",start_color="F2F3F4"); c.font=Font(name="Arial",size=8); continue
                sig_s="" if pv is None or pv>=0.05 else "***" if pv<0.001 else "**" if pv<0.01 else "*"
                c.value=f"{rv:.2f}{sig_s}"; c.font=Font(name="Arial",size=8)
                if rv>=0.5: c.fill=PatternFill("solid",start_color="1A6634"); c.font=Font(name="Arial",size=8,color="FFFFFF",bold=True)
                elif rv>=0.3: c.fill=PatternFill("solid",start_color="A9DFBF")
                elif rv<=-0.5: c.fill=PatternFill("solid",start_color="922B21"); c.font=Font(name="Arial",size=8,color="FFFFFF",bold=True)
                elif rv<=-0.3: c.fill=PatternFill("solid",start_color="F1948A")
            row+=1
    for j,w in enumerate([28,28,28,28,8,14,8,10,10],1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes="A4"

# ================================================================
# SHEET 07 — COMPLIANCE (from synthesis)
# ================================================================
def build_sheet7(wb, synthesis):
    ws=wb.create_sheet("07_Chatbot_Compliance")
    section_title(ws,"Chatbot Compliance — Perceived tone vs requested brief",COLOR_HDR,5,1)
    comp=synthesis.get("compliance",{}) if synthesis else {}
    if not comp: ws.cell(2,1,value="Run 06_synthesis.py first"); return
    pv=comp.get("per_version",{})
    for j,h in enumerate(["Metric","FL_21 (Friendly)","FL_22 (Professional)","Difference"],1):
        hdr(ws.cell(2,j,value=h),bg="2C3E50" if j==1 else COLOR_FL21 if j==2 else COLOR_FL22 if j==3 else "2C3E50",sz=10)
    metrics=[("N respondents","n"),("Compliance score (mean)","compliance_mean"),("SD","compliance_sd"),
             ("% Compliant","pct_compliant"),("% Emojis","pct_emojis"),("% Informal address","pct_informal"),
             ("% Formal address","pct_formal"),("Avg encouragements","avg_encouragements"),
             ("Avg sober phrases","avg_sober"),("Avg friendly words","avg_friendly_words"),("Avg formal words","avg_formal_words")]
    for i,(lbl,key) in enumerate(metrics):
        r=3+i; alt=(i%2==0)
        dat(ws.cell(r,1),lbl,center=False,alt=alt,bold=True)
        v21=pv.get("FL_21",{}).get(key); v22=pv.get("FL_22",{}).get(key)
        dat(ws.cell(r,2),v21,alt=alt,fmt="0.00" if isinstance(v21,float) else None)
        dat(ws.cell(r,3),v22,alt=alt,fmt="0.00" if isinstance(v22,float) else None)
        if isinstance(v21,(int,float)) and isinstance(v22,(int,float)):
            delta_cell(ws.cell(r,4),v21-v22)
        else: dat(ws.cell(r,4),"—",alt=alt)
    tc=comp.get("ttest_compliance",{})
    if tc:
        r=3+len(metrics)+2
        section_title(ws,"T-test: compliance score FL_21 vs FL_22","2C3E50",5,r); r+=1
        write_ttest_table(ws,[tc],r,"2C3E50")
    for j,w in enumerate([32,22,24,16],1): ws.column_dimensions[get_column_letter(j)].width=w

# ================================================================
# SHEET 08 — DROPOUT
# ================================================================
def build_sheet8(wb, drop_data):
    ws=wb.create_sheet("08_Dropout_Analysis")
    section_title(ws,"Dropout Analysis",COLOR_HDR,12,1)
    if not drop_data: ws.cell(2,1,value="Run 08_new_analyses.py first"); return
    row=2
    summary_box(ws,f"Total n={drop_data.get('n_total')} | Proper end: {drop_data.get('n_completers')} ({100-drop_data.get('pct_dropout',0):.1f}%) | Dropouts: {drop_data.get('n_dropouts')} ({drop_data.get('pct_dropout')}%) | FL_21 dropout: {drop_data.get('pct_dropout_fl21')}% | FL_22 dropout: {drop_data.get('pct_dropout_fl22')}%",row,12); row+=2

    # End type table with chi2 per category
    section_title(ws,"End type distribution by tone + significance test per category",C_CONV,12,row); row+=1
    summary_box(ws,"Chi² (or Fisher exact if cells <5) tests whether each end_type proportion differs significantly between FL_21 and FL_22.",row,12); row+=1
    hdrs=["End type","N FL_21","% FL_21","N FL_22","% FL_22","N Total","% Total","Chi²/Fisher","p-value","Sig.","Test"]
    for j,h in enumerate(hdrs,1): hdr(ws.cell(row,j,value=h),bg=C_CONV,sz=9)
    row+=1
    detail=drop_data.get("end_type_detail",[])
    for i,d in enumerate(detail):
        alt=(i%2==0)
        dat(ws.cell(row,1),d.get("category",""),center=False,alt=alt,bold=True)
        dat(ws.cell(row,2),d.get("n_fl21"),alt=alt)
        dat(ws.cell(row,3),d.get("pct_fl21"),alt=alt,fmt="0.0")
        dat(ws.cell(row,4),d.get("n_fl22"),alt=alt)
        dat(ws.cell(row,5),d.get("pct_fl22"),alt=alt,fmt="0.0")
        dat(ws.cell(row,6),d.get("n_total"),alt=alt)
        dat(ws.cell(row,7),d.get("pct_total"),alt=alt,fmt="0.0")
        dat(ws.cell(row,8),d.get("chi2"),alt=alt,fmt="0.0000")
        pval_cell(ws.cell(row,9),d.get("p"))
        dat(ws.cell(row,10),d.get("sig",""),alt=alt,bold=True)
        dat(ws.cell(row,11),d.get("test",""),alt=alt)
        row+=1
    # Global chi2
    chi2=drop_data.get("chi2_by_tone",{})
    ws.cell(row,1,value=f"Global chi² (all categories × tone): χ²={chi2.get('chi2')}  p={chi2.get('p')}  {chi2.get('sig')}").font=Font(italic=True,name="Arial",size=9)
    row+=2

    # Profile comparison dropouts vs completers
    section_title(ws,"Profile comparison — Dropouts vs Completers",C_CONV,12,row); row+=1
    prof=drop_data.get("profile_comparison",[])
    if prof:
        hdrs2=["Variable","Label","Mean dropouts","SD dropouts","Mean completers","SD completers","t","p-value","Cohen's d","Effect size"]
        for j,h in enumerate(hdrs2,1): hdr(ws.cell(row,j,value=h),bg=C_CONV,sz=9)
        row+=1
        for i,r in enumerate(prof):
            alt=(i%2==0)
            dat(ws.cell(row,1),r.get("variable",""),center=False,alt=alt,bold=True)
            dat(ws.cell(row,2),r.get("label",""),center=False,alt=alt)
            dat(ws.cell(row,3),r.get("mean_dropout"),alt=alt,fmt="0.000")
            dat(ws.cell(row,4),r.get("sd_dropout"),alt=alt,fmt="0.000")
            dat(ws.cell(row,5),r.get("mean_completer"),alt=alt,fmt="0.000")
            dat(ws.cell(row,6),r.get("sd_completer"),alt=alt,fmt="0.000")
            dat(ws.cell(row,7),r.get("t"),alt=alt,fmt="0.0000")
            pval_cell(ws.cell(row,8),r.get("p"))
            dat(ws.cell(row,9),r.get("cohens_d"),alt=alt,fmt="0.000")
            dat(ws.cell(row,10),r.get("effect_size",""),alt=alt)
            row+=1
    row+=2

    # E1/E2 by end_type
    section_title(ws,"E1 (Effort) and E2 (Engagement) by end_type — complementary analysis",C_EVAL,12,row); row+=1
    summary_box(ws,"For each end_type: mean_in = participants in that category / mean_out = all others. Significant p means that group scored meaningfully differently on E1 or E2.",row,12); row+=1
    e1e2=drop_data.get("e1_e2_by_endtype",[])
    if e1e2:
        hdrs3=["End type","Variable","Label","N in","Mean in","SD in","N out","Mean out","SD out","t","p-value","Cohen's d","Interpretation"]
        for j,h in enumerate(hdrs3,1): hdr(ws.cell(row,j,value=h),bg=C_EVAL,sz=9)
        row+=1
        for i,r in enumerate(e1e2):
            alt=(i%2==0)
            dat(ws.cell(row,1),r.get("end_type",""),center=False,alt=alt,bold=True)
            dat(ws.cell(row,2),r.get("variable",""),center=False,alt=alt)
            dat(ws.cell(row,3),r.get("label",""),center=False,alt=alt)
            dat(ws.cell(row,4),r.get("n_in"),alt=alt); dat(ws.cell(row,5),r.get("mean_in"),alt=alt,fmt="0.000")
            dat(ws.cell(row,6),r.get("sd_in"),alt=alt,fmt="0.000"); dat(ws.cell(row,7),r.get("n_out"),alt=alt)
            dat(ws.cell(row,8),r.get("mean_out"),alt=alt,fmt="0.000"); dat(ws.cell(row,9),r.get("sd_out"),alt=alt,fmt="0.000")
            dat(ws.cell(row,10),r.get("t"),alt=alt,fmt="0.0000")
            pval_cell(ws.cell(row,11),r.get("p"))
            dat(ws.cell(row,12),r.get("cohens_d"),alt=alt,fmt="0.000")
            dat(ws.cell(row,13),r.get("interpretation",""),center=False,alt=alt)
            row+=1
    for j,w in enumerate([20,12,22,6,10,8,6,10,8,10,14,10,40],1): ws.column_dimensions[get_column_letter(j)].width=w

# ================================================================
# SHEET 09 — TONE COMPARISONS
# ================================================================
def build_sheet9(wb, tone_data):
    ws=wb.create_sheet("09_Tone_Comparisons")
    section_title(ws,"Tone Comparisons — All DVs by Chatbot Tone (FL_21 Friendly vs FL_22 Professional)",COLOR_HDR,14,1)
    if not tone_data: ws.cell(2,1,value="Run 08_new_analyses.py first"); return

    all_t=(tone_data.get("evaluation",[])+tone_data.get("quality_ai",[])+
           tone_data.get("text_metrics",[])+tone_data.get("perception_ai",[])+tone_data.get("conversation",[]))
    sig_t=[r for r in all_t if r and r.get("sig") not in ("ns","n/a",None)]
    act_sig=[r for r in tone_data.get("actionability",[]) if r and r.get("sig") not in ("ns","n/a",None)]
    row=2
    summary_box(ws,f"SUMMARY — {len(sig_t)+len(act_sig)} significant variables out of {len(all_t)+len(tone_data.get('actionability',[]))} tested. Green Δ = FL_21 higher. Red Δ = FL_22 higher.",row,14); row+=2

    for bg,title,key in [
        (C_EVAL,"EVALUATION QUESTIONNAIRE — Participant experience (Likert 1-7)","evaluation"),
        (C_QUALITY,"AI-RATED RESPONSE QUALITY (1-5)","quality_ai"),
        (C_METRICS,"TEXT METRICS — Objective NLP metrics","text_metrics"),
        (C_PSY,"PERCEPTION OF AI — 12 items","perception_ai"),
        (C_CONV,"CONVERSATION & COMPLIANCE","conversation")]:
        section_title(ws,title,bg,14,row); row+=1
        data=tone_data.get(key,[])
        row=write_ttest_table(ws,[r for r in data if r],row,bg)+1

    # Actionability chi2
    section_title(ws,"ACTIONABILITY & CONTENT — Binary indicators (Chi² tests)",C_ACTION,14,row); row+=1
    row=write_chi2_table(ws,[r for r in tone_data.get("actionability",[]) if r],row,C_ACTION)+2

    # Construct composites
    composites=tone_data.get("construct_composites",[])
    if composites:
        section_title(ws,"CONSTRUCT COMPOSITES — Mean per theoretical construct",C_PSY,14,row); row+=1
        summary_box(ws,"Each composite = mean of all items within the construct.",row,14); row+=1
        row=write_ttest_table(ws,composites,row,C_PSY)+2

    # End type table
    section_title(ws,"END TYPE & DROPOUT BY TONE",C_CONV,14,row); row+=1
    detail=tone_data.get("end_type_detail",[])
    if detail:
        hdrs=["End type","N FL_21","% FL_21","N FL_22","% FL_22","N Total","% Total","Chi²/Fisher","p-value","Sig.","Test"]
        for j,h in enumerate(hdrs,1): hdr(ws.cell(row,j,value=h),bg=C_CONV,sz=9)
        row+=1
        for i,d in enumerate(detail):
            alt=(i%2==0)
            dat(ws.cell(row,1),d.get("category",""),center=False,alt=alt,bold=True)
            dat(ws.cell(row,2),d.get("n_fl21"),alt=alt); dat(ws.cell(row,3),d.get("pct_fl21"),alt=alt,fmt="0.0")
            dat(ws.cell(row,4),d.get("n_fl22"),alt=alt); dat(ws.cell(row,5),d.get("pct_fl22"),alt=alt,fmt="0.0")
            dat(ws.cell(row,6),d.get("n_total"),alt=alt); dat(ws.cell(row,7),d.get("pct_total"),alt=alt,fmt="0.0")
            dat(ws.cell(row,8),d.get("chi2"),alt=alt,fmt="0.0000")
            pval_cell(ws.cell(row,9),d.get("p"))
            dat(ws.cell(row,10),d.get("sig",""),alt=alt,bold=True)
            dat(ws.cell(row,11),d.get("test",""),alt=alt)
            row+=1
        et_g=tone_data.get("end_type",{})
        if et_g:
            ws.cell(row,1,value=f"Global chi² end_type × tone: χ²={et_g.get('chi2')}  p={et_g.get('p')}  {et_g.get('sig')}").font=Font(italic=True,name="Arial",size=9)
            row+=1
    # p-value legend
    row+=2
    for bg,fc,txt in [("1A6634","FFFFFF","p<0.001 ***"),("52BE80","FFFFFF","p<0.01 **"),
                       ("F9E79F","7D6608","p<0.05 *"),("FADBD8","922B21","p≥0.05 ns")]:
        ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=5)
        c=ws.cell(row,1,value=txt); c.fill=PatternFill("solid",start_color=bg)
        c.font=Font(name="Arial",size=9,bold=True,color=fc); c.border=THIN; row+=1
    ws.freeze_panes="A3"

# ================================================================
# SHEET 10 — TONE EFFECTS (mediations + regressions, tone as IV)
# ================================================================
def build_sheet10(wb, med_tone, regression_noton):
    ws=wb.create_sheet("10_Tone_Effects")
    section_title(ws,"Tone Effects — All mediations and regressions where Chatbot Tone = IV",COLOR_HDR,14,1)
    summary_box(ws,"Bootstrap=5000, 95% CI. a=IV→M, b=M→DV|IV, c=total IV→DV, c'=direct IV→DV. Coefficient shown with significance stars (e.g. 0.342**). Sig. = CI excludes 0.",2,14)
    if not med_tone: ws.cell(3,1,value="Run 08_new_analyses.py first"); return
    row=4

    # Build flat list of mediation rows grouped by bloc
    bloc_order=["Q1.4 — Tone → Competence → Perception AI",
                "Tone → Engagement/Effort → Quality",
                "Tone → Appreciation → Preference",
                "Tone → Competence → Engagement",
                "Tone → Emotion → Quality",
                "Tone → Emotion → Perception/Evaluation"]
    blocs_seen={}
    for key,item in med_tone.items():
        b=item.get("bloc","Other")
        if b not in blocs_seen: blocs_seen[b]=[]
        blocs_seen[b].append((key,item))

    all_rows=[]
    for b in bloc_order:
        for key,item in blocs_seen.get(b,[]):
            r=med_to_row(key,item,b)
            r["iv_label"]=r["iv_label"].replace("tone","Chatbot tone")
            all_rows.append(r)

    section_title(ws,"BLOC A — MEDIATIONS: Tone → Mediator → Outcome",C_EVAL,14,row); row+=1
    row=write_grand_mediation_table(ws,all_rows,row,C_EVAL,show_bloc=True)

    # Contextual regressions
    row+=2
    section_title(ws,"BLOC B — CONTEXTUAL REGRESSIONS (tone as control variable in quality/preference models)",C_METRICS,14,row); row+=1
    reg_items=[(
        "Quality & Preference predictors",C_METRICS,
        [regression_noton.get("Q2.1_words_quality",{}),
         regression_noton.get("Q2.2_E2_quality",{}),
         regression_noton.get("Q3.2b_E1E2quality_E6",{})])]
    row=write_grand_regression_table(ws,reg_items,row,C_METRICS)
    ws.freeze_panes="A4"

# ================================================================
# SHEET 11 — VARIABLE RELATIONSHIPS (regressions, tone not IV)
# ================================================================
def build_sheet11(wb, regression_noton):
    ws=wb.create_sheet("11_Variable_Relationships")
    section_title(ws,"Variable Relationships — Regressions where Tone is NOT the independent variable",COLOR_HDR,9,1)
    summary_box(ws,"One row per predictor. n and R² shown on first predictor of each model only. Coefficient shown with significance stars.",2,9)
    if not regression_noton: ws.cell(3,1,value="Run 08_new_analyses.py first"); return
    row=4

    reg_groups=[
        ("BLOC 1 — QUALITY PREDICTORS",C_QUALITY,[
            "Q2.1_words_quality","Q2.2_E2_quality","Q2.3_E4_quality",
            "emotion_quality_global","emotion_quality_precision","emotion_quality_examples",
            "emotion_quality_relevance","emotion_quality_richness"]),
        ("BLOC 2 — EMOTION → PSYCHOLOGICAL & EVALUATION",C_PSY,[
            "emotion_Perceived Manipulati_1_num","emotion_Perceived Manipulati_2_num",
            "emotion_Competence_1_num","emotion_Sense of Independenc_1_num",
            "emotion_evaluation_2_num","emotion_evaluation_3_num"]),
        ("BLOC 3 — REUSE INTENTION & CHATBOT PREFERENCE",C_EVAL,[
            "Q3.1_multi_E5","Q3.1b_full_E5","Q3.2_quality_E6","Q3.2b_E1E2quality_E6"]),
        ("BLOC 4 — PERCEPTION AI → CHATBOT APPRECIATION",C_PSY,["Q4.1_psy_E3"]),
        ("BLOC 5 — BREAKPOINT & CONVERSATION",C_CONV,["Q5.2_bkpt_quality"]),
    ]
    reg_items=[]
    for bloc_lbl,bloc_col,keys in reg_groups:
        items=[regression_noton.get(k,{}) for k in keys if regression_noton.get(k)]
        if items: reg_items.append((bloc_lbl,bloc_col,items))
    row=write_grand_regression_table(ws,reg_items,row,C_QUALITY)

    # Q5.3 t-test
    q53=regression_noton.get("Q5.3_bkpt_exists_quality",{})
    if q53:
        row+=2; m=q53.get("model",{})
        section_title(ws,"BLOC 6 — BREAKPOINT EXISTS → QUALITY (t-test)",C_CONV,9,row); row+=1
        for j,h in enumerate(["Group","N","Mean","SD","t","p-value","Cohen's d","Effect size"],1):
            hdr(ws.cell(row,j,value=h),bg=C_CONV,sz=9)
        row+=1
        dat(ws.cell(row,1),"No breakpoint (0)",center=False,bold=True)
        dat(ws.cell(row,2),m.get("n_no_breakpoint")); dat(ws.cell(row,3),m.get("mean_no_bkpt"),fmt="0.000")
        dat(ws.cell(row,4),m.get("sd_no_bkpt"),fmt="0.000"); dat(ws.cell(row,5),m.get("t"),fmt="0.0000")
        pval_cell(ws.cell(row,6),m.get("p"))
        dat(ws.cell(row,7),m.get("cohens_d"),fmt="0.000"); dat(ws.cell(row,8),m.get("effect_size","")); row+=1
        dat(ws.cell(row,1),"Breakpoint detected (1)",center=False,bold=True,alt=True)
        dat(ws.cell(row,2),m.get("n_breakpoint"),alt=True); dat(ws.cell(row,3),m.get("mean_bkpt"),alt=True,fmt="0.000")
        dat(ws.cell(row,4),m.get("sd_bkpt"),alt=True,fmt="0.000")
    ws.freeze_panes="A4"

# ================================================================
# SHEET 12 — MEDIATIONS WITHOUT TONE
# ================================================================
def build_sheet12(wb, med_noton):
    ws=wb.create_sheet("12_Mediation_Moderation")
    section_title(ws,"Mediations — Variable interactions where Tone is NOT the independent variable",COLOR_HDR,14,1)
    summary_box(ws,"Bootstrap=5000, 95% CI. One row = one model. Significant = CI excludes 0. Coefficient shown with stars. Grouped by research bloc.",2,14)
    if not med_noton: ws.cell(3,1,value="Run 08_new_analyses.py first"); return
    row=4

    # Single models
    single_models=[]
    for key,bloc,lbl in [
        ("Q3.3_E4_E3_E6","E4→E3→E6 (Appreciation pathway)","E4 (Utility) → E3 (Appreciation) → E6 (Chatbot preference)"),
        ("emotion_E2_quality","Emotion mediations","Emotion expressed → E2 (Engagement) → Quality global"),
        ("emotion_E3_E6","Emotion mediations","Emotion expressed → E3 (Appreciation) → E6 (Chatbot preference)")]:
        item=med_noton.get(key,{})
        r=med_to_row(key,item,bloc); r["label"]=lbl; single_models.append(r)

    # Q2.6 — 12 models
    q26=med_noton.get("Q2.6_psy_E2_quality",{}).get("models",[])
    q26_rows=[{"bloc":"Q2.6 — Perception AI → E2 → Quality",
               "label":f"{m.get('iv_label','')} → E2 → Quality","iv_label":m.get("iv_label",""),
               "mediator_label":"E2 Engagement felt","dv_label":"Quality global",
               **({k:m["result"].get(k,{}) for k in ["path_a","path_b","path_c","path_cp","indirect"]} if not m.get("result",{}).get("error") else {"path_a":{},"path_b":{},"path_c":{},"path_cp":{},"indirect":{}}),
               "n":m.get("result",{}).get("n"),"mediation_type":m.get("result",{}).get("mediation_type","Error")} for m in q26]

    # Q4.4 — 12 models
    q44=med_noton.get("Q4.4_psy_E3_E6",{}).get("models",[])
    q44_rows=[{"bloc":"Q4.4 — Perception AI → E3 → E6 Preference",
               "label":f"{m.get('iv_label','')} → E3 → E6","iv_label":m.get("iv_label",""),
               "mediator_label":"E3 Chatbot appreciation","dv_label":"E6 Chatbot preference",
               **({k:m["result"].get(k,{}) for k in ["path_a","path_b","path_c","path_cp","indirect"]} if not m.get("result",{}).get("error") else {"path_a":{},"path_b":{},"path_c":{},"path_cp":{},"indirect":{}}),
               "n":m.get("result",{}).get("n"),"mediation_type":m.get("result",{}).get("mediation_type","Error")} for m in q44]

    all_rows=single_models+q26_rows+q44_rows
    section_title(ws,"ALL MEDIATIONS (Tone not IV) — One row per model",C_CONV,14,row); row+=1
    row=write_grand_mediation_table(ws,all_rows,row,"2C3E50",show_bloc=True)
    ws.freeze_panes="A4"

# ================================================================
# SHEET 13 — QUALITY PROGRESSION
# ================================================================
def build_sheet13(wb, df_prog, synthesis):
    ws=wb.create_sheet("13_Quality_Progression")
    section_title(ws,"Quality Score per Conversation Turn — FL_21 vs FL_22",COLOR_HDR,9,1)
    summary_box(ws,"Quality 1-5 rated per turn by GPT-4o-mini. Only turns with N≥8 per condition plotted.",2,9)
    prog=synthesis.get("progression",{}) if synthesis else {}
    if not prog and not df_prog.empty:
        for _,r in df_prog.iterrows():
            ver=r.get("version",""); t=str(int(r.get("turn",0)))
            prog.setdefault(ver,{}).setdefault(t,{"scores":[]})["scores"].append(r.get("quality_score"))
        for ver in prog:
            for t in prog[ver]:
                sc=[s for s in prog[ver][t].get("scores",[]) if s is not None]
                prog[ver][t]={"mean":round(float(np.mean(sc)),3) if sc else None,
                              "sd":round(float(np.std(sc,ddof=1)),3) if len(sc)>1 else None,"n":len(sc)}
    all_turns=sorted(set(int(t) for v in prog.values() for t in v.keys())) if prog else []
    row=4
    for j,h in enumerate(["Turn","Mean FL_21","SD FL_21","N FL_21","Mean FL_22","SD FL_22","N FL_22","Δ(21-22)"],1):
        hdr(ws.cell(row,j,value=h),bg=COLOR_FL21 if "21" in h else COLOR_FL22 if "22" in h else "2C3E50",sz=9)
    row+=1
    for t in all_turns:
        ts=str(t); alt=(t%2==0)
        d21=prog.get("FL_21",{}).get(ts,{}); d22=prog.get("FL_22",{}).get(ts,{})
        m21=d21.get("mean"); m22=d22.get("mean")
        dat(ws.cell(row,1),t,alt=alt,bold=True)
        dat(ws.cell(row,2),m21,alt=alt,fmt="0.000"); dat(ws.cell(row,3),d21.get("sd"),alt=alt,fmt="0.000")
        dat(ws.cell(row,4),d21.get("n"),alt=alt); dat(ws.cell(row,5),m22,alt=alt,fmt="0.000")
        dat(ws.cell(row,6),d22.get("sd"),alt=alt,fmt="0.000"); dat(ws.cell(row,7),d22.get("n"),alt=alt)
        if m21 and m22: delta_cell(ws.cell(row,8),m21-m22)
        else: ws.cell(row,8,value="n/a").border=THIN
        row+=1
    if prog and all_turns:
        fig,ax=plt.subplots(figsize=(12,5))
        for ver,col,lbl in [("FL_21",COLOR_FL21,"FL_21 Friendly"),("FL_22",COLOR_FL22,"FL_22 Professional")]:
            dv=prog.get(ver,{})
            tv=[int(t) for t in dv if dv[t].get("n",0)>=8]
            mv=[dv[str(t)].get("mean") for t in tv if dv.get(str(t),{}).get("mean")]
            valid=[(t,m) for t,m in zip(tv,mv) if m is not None]
            if valid:
                tv2,mv2=zip(*valid)
                ax.plot(tv2,mv2,marker="o",color=f"#{col}",linewidth=2,label=lbl)
        ax.set_xlabel("Conversation turn"); ax.set_ylabel("Avg quality score (1-5)")
        ax.set_title("Quality score per turn",fontsize=12,fontweight="bold")
        ax.legend(); ax.grid(alpha=0.3); ax.set_ylim(1,5); plt.tight_layout()
        fig_path=OUTPUT_DIR/"fig_progression.png"; fig.savefig(fig_path,dpi=150,bbox_inches="tight"); plt.close(fig)
        insert_png(ws,fig_path,"J4",22,9)

# ================================================================
# SHEET 14 — WORD FREQUENCIES
# ================================================================
def build_sheet14(wb, wordfreq):
    ws=wb.create_sheet("14_Word_Frequencies")
    section_title(ws,"Word Frequencies & TF-IDF — Participants and Chatbot responses",COLOR_HDR,9,1)
    if not wordfreq: ws.cell(2,1,value="Run 04_wordfreq.py first"); return
    tw_m21=wordfreq.get("top_words_participants_fl21",[]); tw_m22=wordfreq.get("top_words_participants_fl22",[])
    tw_r21=wordfreq.get("top_words_chatbot_fl21",[]); tw_r22=wordfreq.get("top_words_chatbot_fl22",[])
    row=3
    for j,h in enumerate(["Rank","Word FL_21 (Part.)","Freq","Word FL_22 (Part.)","Freq",
                           "Word Chatbot FL_21","Freq","Word Chatbot FL_22","Freq"],1):
        hdr(ws.cell(row,j,value=h),bg=COLOR_FL21 if "21" in h else COLOR_FL22 if "22" in h else "2C3E50",sz=9)
    row+=1
    for i in range(min(40,max(len(tw_m21),len(tw_m22),len(tw_r21),len(tw_r22),1))):
        alt=(i%2==0); dat(ws.cell(row,1),i+1,alt=alt,bold=True)
        if i<len(tw_m21): dat(ws.cell(row,2),tw_m21[i][0],center=False,alt=alt); dat(ws.cell(row,3),tw_m21[i][1],alt=alt)
        if i<len(tw_m22): dat(ws.cell(row,4),tw_m22[i][0],center=False,alt=alt); dat(ws.cell(row,5),tw_m22[i][1],alt=alt)
        if i<len(tw_r21): dat(ws.cell(row,6),tw_r21[i][0],center=False,alt=alt); dat(ws.cell(row,7),tw_r21[i][1],alt=alt)
        if i<len(tw_r22): dat(ws.cell(row,8),tw_r22[i][0],center=False,alt=alt); dat(ws.cell(row,9),tw_r22[i][1],alt=alt)
        row+=1
    row+=2
    section_title(ws,"Distinctive words (TF-IDF)","C0392B",9,row); row+=1
    for j,h in enumerate(["Rank","Distinctive FL_21 (Part.)","Δ TF-IDF","Distinctive FL_22 (Part.)","Δ TF-IDF",
                           "Distinctive Chatbot FL_21","Δ TF-IDF","Distinctive Chatbot FL_22","Δ TF-IDF"],1):
        hdr(ws.cell(row,j,value=h),bg="2C3E50",sz=9)
    row+=1
    for i in range(20):
        alt=(i%2==0); dat(ws.cell(row,1),i+1,alt=alt,bold=True)
        for ci,key in [(2,"tfidf_participants_fl21"),(4,"tfidf_participants_fl22"),(6,"tfidf_chatbot_fl21"),(8,"tfidf_chatbot_fl22")]:
            tw=wordfreq.get(key,[])
            if i<len(tw): dat(ws.cell(row,ci),tw[i][0],center=False,alt=alt); dat(ws.cell(row,ci+1),tw[i][2],alt=alt,fmt="0.0000")
        row+=1
    for fname,anchor in [("fig_wc_part_fl21.png","K3"),("fig_wc_part_fl22.png","K18"),
                          ("fig_wc_bot_fl21.png","K33"),("fig_wc_bot_fl22.png","K48")]:
        insert_png(ws,OUTPUT_DIR/fname,anchor,20,8)
    for j,w in enumerate([6,22,10,22,10,22,10,22,10],1): ws.column_dimensions[get_column_letter(j)].width=w

# ================================================================
# MAIN
# ================================================================
def run():
    print("\nBuilding Excel (14 sheets + TOC)...")
    d=load_all()
    wb=Workbook()
    print("  Sheet 00 — TOC"); build_sheet0(wb)
    print("  Sheet 01 — Raw data"); build_raw_sheet(wb,d["df_clean"],"Raw Cleaned Data","01_Raw_Data_Cleaned",COLOR_HDR)
    print("  Sheet 02 — FL_21"); build_raw_sheet(wb,d["df_fl21"],"FL_21 — Friendly tone","02_FL21_Friendly",COLOR_FL21)
    print("  Sheet 03 — FL_22"); build_raw_sheet(wb,d["df_fl22"],"FL_22 — Professional tone","03_FL22_Professional",COLOR_FL22)
    print("  Sheet 04 — Scale coding"); build_sheet4(wb,d["df_coding"])
    print("  Sheet 05 — Variable dictionary"); build_sheet5(wb)
    print("  Sheet 06 — Correlations"); build_sheet6(wb,d["correlations"])
    print("  Sheet 07 — Compliance"); build_sheet7(wb,d["synthesis"])
    print("  Sheet 08 — Dropout"); build_sheet8(wb,d["dropout"])
    print("  Sheet 09 — Tone comparisons"); build_sheet9(wb,d["tone_comparisons"])
    print("  Sheet 10 — Tone effects"); build_sheet10(wb,d["mediation_tone"],d["regression_noton"])
    print("  Sheet 11 — Variable relationships"); build_sheet11(wb,d["regression_noton"])
    print("  Sheet 12 — Mediations"); build_sheet12(wb,d["mediation_noton"])
    print("  Sheet 13 — Quality progression"); build_sheet13(wb,d["df_prog"],d["synthesis"])
    print("  Sheet 14 — Word frequencies"); build_sheet14(wb,d["wordfreq"])
    wb.save(OUTPUT_FILE)
    print(f"\nExcel saved: {OUTPUT_FILE}")
    return str(OUTPUT_FILE)

if __name__=="__main__":
    run()
