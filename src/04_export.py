bash

cat > /home/claude/thesis_project/src/04_export.py << 'PYEOF'
"""
04_export.py — Génère le fichier Excel final avec toutes les feuilles.
Lance aussi les étapes 01, 02, 03 si les JSON intermédiaires n'existent pas.
"""
import sys, json, io
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy import stats
from scipy.stats import chi2_contingency
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from config import OUTPUT_DIR, OUTPUT_FILE, MSG_COLS, RESP_COLS, PALETTE, COLOR_FL21, COLOR_FL22, COLOR_HDR, COLOR_ALT
from loader import load_clean_data

# ════════════════════════════════════════════════════════
# HELPERS OPENPYXL
# ════════════════════════════════════════════════════════
THIN = Border(
    left=Side(style="thin",color="BFBFBF"), right=Side(style="thin",color="BFBFBF"),
    top=Side(style="thin",color="BFBFBF"),  bottom=Side(style="thin",color="BFBFBF"),
)
def hdr(c, bg=COLOR_HDR, fc="FFFFFF", sz=10, bold=True):
    c.font      = Font(bold=bold, color=fc, name="Arial", size=sz)
    c.fill      = PatternFill("solid", start_color=bg.lstrip("#"))
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = THIN

def dat(c, val, center=True, alt=False, bold=False, fc="000000", fmt=None):
    v = None if (isinstance(val, float) and np.isnan(val)) else val
    if isinstance(v, bool): v = "Oui" if v else "Non"
    c.value     = v
    c.font      = Font(name="Arial", size=9, bold=bold, color=fc)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    c.border    = THIN
    if alt: c.fill = PatternFill("solid", start_color=COLOR_ALT.lstrip("#"))
    if fmt: c.number_format = fmt

def pval(c, p):
    c.border    = THIN
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font      = Font(name="Arial", size=9)
    if p is None or (isinstance(p, float) and np.isnan(p)):
        c.value = "n/a"; return
    if   p < 0.001: c.value=f"{p:.4f} ***"; c.fill=PatternFill("solid",start_color="1A6634"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif p < 0.01:  c.value=f"{p:.4f} **";  c.fill=PatternFill("solid",start_color="52BE80"); c.font=Font(name="Arial",size=9,bold=True,color="FFFFFF")
    elif p < 0.05:  c.value=f"{p:.4f} *";   c.fill=PatternFill("solid",start_color="F9E79F"); c.font=Font(name="Arial",size=9,bold=True,color="7D6608")
    else:           c.value=f"{p:.4f} ns";  c.fill=PatternFill("solid",start_color="FADBD8"); c.font=Font(name="Arial",size=9,color="922B21")

def title_row(ws, text, bg, ncols, row=1, sz=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=sz)
    c.fill      = PatternFill("solid", start_color=bg.lstrip("#"))
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26

def insert_png(ws, png_path, anchor, w_cm=18, h_cm=9):
    if not Path(png_path).exists(): return
    img = XLImage(str(png_path))
    img.width  = int(w_cm * 37.795)
    img.height = int(h_cm * 37.795)
    ws.add_image(img, anchor)

def write_ttest_table(ws, ttest_list, start_row=3, bg_header="2C3E50"):
    hdrs = ["Métrique","Moy. FL_21","SD FL_21","Moy. FL_22","SD FL_22","t / chi2","p-value","Sig.","Cohen's d"]
    for j, h in enumerate(hdrs, 1):
        hdr(ws.cell(row=start_row, column=j, value=h), bg=f"#{bg_header}", sz=9)
    ws.row_dimensions[start_row].height = 25
    for i, row_t in enumerate(ttest_list):
        r = start_row + 1 + i
        alt = (i % 2 == 0)
        dat(ws.cell(row=r, column=1), row_t.get("label",""), center=False, alt=alt, bold=True)
        dat(ws.cell(row=r, column=2), row_t.get("mean_fl21"), alt=alt, fmt="0.000")
        dat(ws.cell(row=r, column=3), row_t.get("sd_fl21"), alt=alt, fmt="0.000")
        dat(ws.cell(row=r, column=4), row_t.get("mean_fl22"), alt=alt, fmt="0.000")
        dat(ws.cell(row=r, column=5), row_t.get("sd_fl22"), alt=alt, fmt="0.000")
        dat(ws.cell(row=r, column=6), row_t.get("t"), alt=alt, fmt="0.0000")
        pval(ws.cell(row=r, column=7), row_t.get("p"))
        dat(ws.cell(row=r, column=8), row_t.get("sig",""), alt=alt, bold=True)
        dat(ws.cell(row=r, column=9), row_t.get("d"), alt=alt, fmt="0.000")
    widths = [32,11,10,11,10,10,14,8,10]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return start_row + 1 + len(ttest_list)

# ════════════════════════════════════════════════════════
# FIGURES
# ════════════════════════════════════════════════════════
def fig_bytes(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf

def fig_turns_distribution(df_clean):
    fig, axes = plt.subplots(1, 2, figsize=(14, 5))
    fig.suptitle("Distribution des tours de conversation — FL_21 vs FL_22", fontsize=13, fontweight="bold")
    for ax, version, color in zip(axes, ["FL_21","FL_22"], PALETTE):
        grp  = df_clean[df_clean["FL_13_DO"]==version]
        dist = grp["n_turns"].value_counts().sort_index()
        bars = ax.bar(dist.index, dist.values, color=color, alpha=0.82, edgecolor="white", width=0.7)
        ax.axvline(grp["n_turns"].mean(), color="black", linestyle="--", linewidth=2,
                   label=f"Moy. = {grp['n_turns'].mean():.1f}")
        label = "Friendly" if version=="FL_21" else "Pro"
        ax.set_title(f"{version} — {label}", color=color, fontweight="bold")
        ax.set_xlabel("Nombre de tours"); ax.set_ylabel("N participants")
        ax.legend(); ax.grid(axis="y", alpha=0.3)
        for bar in bars:
            ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.1,
                    str(int(bar.get_height())), ha="center", fontsize=9)
    plt.tight_layout()
    return fig

def fig_metrics_comparison(ttest_list, title):
    labels = [r["label"] for r in ttest_list]
    m21    = [r.get("mean_fl21", 0) for r in ttest_list]
    m22    = [r.get("mean_fl22", 0) for r in ttest_list]
    sigs   = [r.get("sig","ns") for r in ttest_list]
    x = np.arange(len(labels)); w = 0.35
    fig, ax = plt.subplots(figsize=(max(10, len(labels)*1.5), 6))
    b1 = ax.bar(x-w/2, m21, w, label="FL_21 Friendly", color=COLOR_FL21, alpha=0.82)
    b2 = ax.bar(x+w/2, m22, w, label="FL_22 Pro",      color=COLOR_FL22, alpha=0.82)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right", fontsize=9)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.legend(); ax.grid(axis="y", alpha=0.3)
    for xi, sig in zip(x, sigs):
        if sig != "ns":
            ymax = max(m21[xi] if xi<len(m21) else 0, m22[xi] if xi<len(m22) else 0)
            ax.text(xi, ymax*1.05, sig, ha="center", fontsize=11, color="crimson", fontweight="bold")
    plt.tight_layout()
    return fig

# ════════════════════════════════════════════════════════
# FLATTEN AI RESULTS
# ════════════════════════════════════════════════════════
def flatten_ai(record):
    rid  = record.get("respondent_id","")
    ver  = record.get("version","")
    n_t  = record.get("n_turns", np.nan)
    part = record.get("participant", {})
    bot  = record.get("chatbot", {})
    qr   = part.get("qualite_reponse", {})
    act  = part.get("actionabilite", {})
    eng  = part.get("profil_participant", {})
    cnt  = part.get("contenu", {})
    ton  = bot.get("ton_perçu", {})
    stq  = bot.get("style_questions", {})
    mrk  = bot.get("marqueurs_ton", {})
    return {
        "respondent_id":    rid,
        "version":          ver,
        "n_turns":          n_t,
        # Qualité participant
        "qualite_score":    qr.get("score"),
        "qualite_label":    qr.get("label",""),
        "qualite_just":     qr.get("justification",""),
        # Actionabilité
        "action_score":     act.get("score"),
        "action_pb_concret":act.get("contient_probleme_concret"),
        "action_conseil":   act.get("contient_conseil_applicable"),
        "action_usecase":   act.get("contient_use_case_précis"),
        "action_just":      act.get("justification",""),
        # Engagement
        "engagement":       eng.get("engagement",""),
        "elaboration":      eng.get("elaboration",""),
        "coherence":        eng.get("coherence",""),
        "expertise":        eng.get("expertise_perçue",""),
        # Contenu
        "has_opinion":      cnt.get("opinion_personnelle"),
        "has_frustration":  cnt.get("frustration_exprimee"),
        "has_suggestion":   cnt.get("suggestion_concrete"),
        "has_experience":   cnt.get("experience_vecue"),
        "has_competitor":   cnt.get("comparaison_concurrents"),
        "has_feature_req":  cnt.get("feature_request"),
        # Synthèse participant
        "themes":           " | ".join(part.get("themes",[])),
        "sentiment":        part.get("sentiment_dominant",""),
        "langue":           part.get("langue",""),
        "abandon":          part.get("abandon_premature"),
        "resume":           part.get("resume",""),
        "verbatim":         part.get("verbatim_cle",""),
        # Chatbot
        "bot_chaleur":      ton.get("score_chaleur"),
        "bot_formalisme":   ton.get("score_formalisme"),
        "bot_empathie":     ton.get("score_empathie"),
        "bot_ton_label":    ton.get("label_dominant",""),
        "bot_pct_ouvertes": stq.get("ouvertes_pct"),
        "bot_relances":     stq.get("relances_personnalisees"),
        "bot_encouragement":stq.get("formules_encouragement"),
        "bot_pression":     stq.get("pression_ressentie",""),
        "bot_friendly_markers": " | ".join(mrk.get("friendly_markers",[])),
        "bot_pro_markers":      " | ".join(mrk.get("pro_markers",[])),
        "bot_coherence_ton": bot.get("coherence_ton",""),
        "bot_resume_style":  bot.get("resume_style",""),
        "ok": part.get("classification_ok", False),
    }

def ttest_binary_ai(df_ai, col, label):
    """t-test ou chi2 sur colonne bool/int de df_ai."""
    g21 = df_ai[df_ai["version"]=="FL_21"][col].apply(lambda x: int(x) if isinstance(x,bool) else (x or 0)).dropna().values
    g22 = df_ai[df_ai["version"]=="FL_22"][col].apply(lambda x: int(x) if isinstance(x,bool) else (x or 0)).dropna().values
    if len(g21)<2 or len(g22)<2:
        return {"label":label,"mean_fl21":float(np.mean(g21)) if len(g21) else np.nan,
                "mean_fl22":float(np.mean(g22)) if len(g22) else np.nan,"t":None,"p":None,"sig":"n/a","d":None}
    t_s, p_v = stats.ttest_ind(g21, g22)
    sig  = "***" if p_v<0.001 else "**" if p_v<0.01 else "*" if p_v<0.05 else "ns"
    pool = np.sqrt(((len(g21)-1)*np.std(g21,ddof=1)**2+(len(g22)-1)*np.std(g22,ddof=1)**2)/(len(g21)+len(g22)-2))
    d    = (np.mean(g21)-np.mean(g22))/pool if pool>0 else 0
    return {"label":label,"mean_fl21":round(float(np.mean(g21)),3),"sd_fl21":round(float(np.std(g21,ddof=1)),3),
            "mean_fl22":round(float(np.mean(g22)),3),"sd_fl22":round(float(np.std(g22,ddof=1)),3),
            "t":round(float(t_s),4),"p":round(float(p_v),4),"sig":sig,"d":round(float(d),3)}

# ════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════
def run():
    # ── Chargement données ────────────────────────────────
    df_clean, _, _ = load_clean_data()

    # ── Lancement étapes précédentes si nécessaire ────────
    metrics_path  = OUTPUT_DIR / "metrics.json"
    wordfreq_path = OUTPUT_DIR / "wordfreq.json"
    ai_path       = OUTPUT_DIR / "ai_results.json"

    if not metrics_path.exists():
        print("Étape 1 — Calcul des métriques...")
        import importlib, os
        sys.path.insert(0, str(Path(__file__).parent))
        import importlib.util
        spec = importlib.util.spec_from_file_location("m01", Path(__file__).parent/"01_metrics.py")
        m01 = importlib.util.module_from_spec(spec); spec.loader.exec_module(m01); m01.run()

    if not wordfreq_path.exists():
        print("Étape 2 — Fréquences de mots...")
        spec = importlib.util.spec_from_file_location("m02", Path(__file__).parent/"02_wordfreq.py")
        m02 = importlib.util.module_from_spec(spec); spec.loader.exec_module(m02); m02.run()

    if not ai_path.exists():
        print("Étape 3 — Classification IA (appels API)...")
        spec = importlib.util.spec_from_file_location("m03", Path(__file__).parent/"03_api_analysis.py")
        m03 = importlib.util.module_from_spec(spec); spec.loader.exec_module(m03); m03.run()

    # ── Lecture des résultats ─────────────────────────────
    with open(metrics_path,  encoding="utf-8") as f: metrics  = json.load(f)
    with open(wordfreq_path, encoding="utf-8") as f: wordfreq = json.load(f)

    df_msg  = pd.read_json(OUTPUT_DIR / "df_msg.json")
    df_resp = pd.read_json(OUTPUT_DIR / "df_resp.json")
    df_agg  = pd.read_json(OUTPUT_DIR / "df_agg.json")

    # ── Résultats IA ──────────────────────────────────────
    df_ai = pd.DataFrame()
    if ai_path.exists():
        with open(ai_path, encoding="utf-8") as f:
            ai_raw = json.load(f)
        df_ai = pd.DataFrame([flatten_ai(v) for v in ai_raw.values() if v.get("participant",{}).get("classification_ok")])
    else:
        print("⚠️  Pas de résultats IA (ai_results.json absent). Lancez 03_api_analysis.py.")

    # ── Figures génériques ────────────────────────────────
    fig_turns = fig_turns_distribution(df_clean)
    fig_turns.savefig(OUTPUT_DIR/"fig_turns.png", dpi=150, bbox_inches="tight"); plt.close(fig_turns)

    fig_m = fig_metrics_comparison(metrics["msg_per_message"],
                                   "Métriques de longueur — Messages participants (FL_21 vs FL_22)")
    fig_m.savefig(OUTPUT_DIR/"fig_metrics_msg.png", dpi=150, bbox_inches="tight"); plt.close(fig_m)

    fig_r = fig_metrics_comparison(metrics["bot_per_message"],
                                   "Métriques de longueur — Réponses chatbot (FL_21 vs FL_22)")
    fig_r.savefig(OUTPUT_DIR/"fig_metrics_bot.png", dpi=150, bbox_inches="tight"); plt.close(fig_r)

    if not df_ai.empty:
        ai_content_ttests = [
            ttest_binary_ai(df_ai, col, lbl) for col, lbl in [
                ("qualite_score",   "Score qualité réponses (1-5)"),
                ("action_score",    "Score actionabilité (1-5)"),
                ("action_pb_concret","Problème concret exprimé"),
                ("action_conseil",   "Conseil applicable"),
                ("action_usecase",   "Use case précis"),
                ("has_opinion",      "Opinion personnelle"),
                ("has_frustration",  "Frustration exprimée"),
                ("has_suggestion",   "Suggestion concrète"),
                ("has_experience",   "Expérience vécue"),
                ("has_feature_req",  "Feature request"),
                ("has_competitor",   "Référence concurrents"),
                ("bot_chaleur",      "Ton chatbot — chaleur (1-5)"),
                ("bot_formalisme",   "Ton chatbot — formalisme (1-5)"),
                ("bot_empathie",     "Ton chatbot — empathie (1-5)"),
                ("bot_pct_ouvertes", "% questions ouvertes chatbot"),
            ]
        ]
        fig_ai = fig_metrics_comparison(
            [r for r in ai_content_ttests if r.get("mean_fl21") is not None][:8],
            "Profil des réponses (analyse IA) — FL_21 vs FL_22")
        fig_ai.savefig(OUTPUT_DIR/"fig_ai_content.png", dpi=150, bbox_inches="tight"); plt.close(fig_ai)

    # ════════════════════════════════════════════════════
    # EXCEL
    # ════════════════════════════════════════════════════
    wb = Workbook()

    # ──────────────────────────────────────────────────
    # FEUILLE 1 — Tours de conversation
    # ──────────────────────────────────────────────────
    ws1 = wb.active; ws1.title = "1_Tours_Conversation"
    title_row(ws1, "Analyse des tours de conversation — FL_21 (Friendly) vs FL_22 (Pro)", "#1F4E79", 9)

    turns_data = metrics["turns"]
    # Tableau résumé
    summary = [
        {"Indicateur":"Moy. tours FL_21","Valeur": turns_data["ttest"]["mean_fl21"]},
        {"Indicateur":"Moy. tours FL_22","Valeur": turns_data["ttest"]["mean_fl22"]},
        {"Indicateur":"Test t (p-value)","Valeur": turns_data["ttest"]["p"]},
        {"Indicateur":"Signif. t-test",  "Valeur": turns_data["ttest"]["sig"]},
        {"Indicateur":"% fin propre FL_21","Valeur":f"{turns_data['pct_ended_fl21']}%"},
        {"Indicateur":"% fin propre FL_22","Valeur":f"{turns_data['pct_ended_fl22']}%"},
        {"Indicateur":"Chi2 p-value (fin propre)","Valeur": turns_data["chi2_end_proper"]["p"]},
    ]
    for j, h in enumerate(["Indicateur","Valeur"], 1):
        hdr(ws1.cell(row=2, column=j, value=h), bg="#2C3E50", sz=9)
    for i, row_s in enumerate(summary):
        alt = (i%2==0)
        dat(ws1.cell(row=3+i, column=1), row_s["Indicateur"], center=False, alt=alt, bold=True)
        c = ws1.cell(row=3+i, column=2)
        v = row_s["Valeur"]
        if isinstance(v, float) and str(v).replace("0.","").replace(".","").isdigit():
            pval(c, v) if "p-value" in row_s["Indicateur"] or "p" == str(v)[:1] else dat(c, v, alt=alt, fmt="0.000")
        else:
            dat(c, v, alt=alt)

    # Distribution détaillée
    row_off = 12
    ws1.cell(row=row_off, column=1, value="Distribution détaillée :").font = Font(bold=True, name="Arial", size=10)
    row_off += 1
    for j, h in enumerate(["N tours","FL_21 (n)","FL_21 (%)","FL_22 (n)","FL_22 (%)"], 1):
        hdr(ws1.cell(row=row_off, column=j, value=h),
            bg="#1A6634" if "21" in h else "#7B3F00" if "22" in h else "#2C3E50", sz=9)
    row_off += 1
    grp21 = df_clean[df_clean["FL_13_DO"]=="FL_21"]; n21=len(grp21)
    grp22 = df_clean[df_clean["FL_13_DO"]=="FL_22"]; n22=len(grp22)
    all_t  = sorted(set(list(turns_data["distribution"]["FL_21"].keys()) + list(turns_data["distribution"]["FL_22"].keys())), key=int)
    for i, t in enumerate(all_t):
        c21 = int(turns_data["distribution"]["FL_21"].get(str(t),0))
        c22 = int(turns_data["distribution"]["FL_22"].get(str(t),0))
        alt = (i%2==0)
        dat(ws1.cell(row=row_off+i, column=1), int(t), alt=alt, bold=True)
        dat(ws1.cell(row=row_off+i, column=2), c21, alt=alt)
        dat(ws1.cell(row=row_off+i, column=3), round(c21/n21*100,1), alt=alt, fmt="0.0")
        dat(ws1.cell(row=row_off+i, column=4), c22, alt=alt)
        dat(ws1.cell(row=row_off+i, column=5), round(c22/n22*100,1), alt=alt, fmt="0.0")

    insert_png(ws1, OUTPUT_DIR/"fig_turns.png", "G2", 22, 10)
    for j, w in enumerate([30,12,12,12,12,5,5,5,5,5], 1):
        ws1.column_dimensions[get_column_letter(j)].width = w
    print("Feuille 1 — Tours OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 2 — Métriques participants (longueur etc.)
    # ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("2_Métriques_Participants")
    title_row(ws2, "Métriques de longueur & structure — Messages participants", "#2E75B6", 9)
    next_r = write_ttest_table(ws2, metrics["msg_per_message"], start_row=2)
    ws2.cell(row=next_r+1, column=1, value="Tests t par participant (agrégés) :").font=Font(bold=True,name="Arial",size=10)
    write_ttest_table(ws2, metrics["msg_per_participant"], start_row=next_r+2)
    insert_png(ws2, OUTPUT_DIR/"fig_metrics_msg.png", "K2", 22, 9)
    print("Feuille 2 — Métriques participants OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 3 — Métriques chatbot
    # ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("3_Métriques_Chatbot")
    title_row(ws3, "Métriques de longueur & structure — Réponses chatbot", "#7B3F00", 9)
    write_ttest_table(ws3, metrics["bot_per_message"], start_row=2)
    insert_png(ws3, OUTPUT_DIR/"fig_metrics_bot.png", "K2", 22, 9)
    print("Feuille 3 — Métriques chatbot OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 4 — Top mots participants
    # ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("4_Mots_Participants")
    title_row(ws4, "Fréquences de mots — Participants (après stopwords)", "#1A6634", 8)
    max_r = max(len(wordfreq["top_words_participants_fl21"]), len(wordfreq["top_words_participants_fl22"]))
    for j, h in enumerate(["Rang","Mot FL_21","Freq","Mot FL_22","Freq","En commun"], 1):
        hdr(ws4.cell(row=2,column=j,value=h),
            bg="#1A6634" if "21" in h else "#7B3F00" if "22" in h else "#2C3E50", sz=9)
    m22_words = {w for w,_ in wordfreq["top_words_participants_fl22"]}
    for i in range(max_r):
        alt = (i%2==0)
        dat(ws4.cell(row=3+i,column=1), i+1, alt=alt, bold=True)
        if i < len(wordfreq["top_words_participants_fl21"]):
            w21, c21 = wordfreq["top_words_participants_fl21"][i]
            dat(ws4.cell(row=3+i,column=2), w21, center=False, alt=alt)
            dat(ws4.cell(row=3+i,column=3), c21, alt=alt)
            dat(ws4.cell(row=3+i,column=6), "✓" if w21 in m22_words else "", alt=alt)
        if i < len(wordfreq["top_words_participants_fl22"]):
            w22, c22 = wordfreq["top_words_participants_fl22"][i]
            dat(ws4.cell(row=3+i,column=4), w22, center=False, alt=alt)
            dat(ws4.cell(row=3+i,column=5), c22, alt=alt)
    # TF-IDF
    tfidf_row = max_r + 5
    title_row(ws4, "Mots distinctifs TF-IDF — Participants (mots qui différencient vraiment les deux groupes)", "#C0392B", 8, row=tfidf_row)
    for j, h in enumerate(["Rang","Mot distinctif FL_21","Δ TF-IDF","Mot distinctif FL_22","Δ TF-IDF"], 1):
        hdr(ws4.cell(row=tfidf_row+1,column=j,value=h),
            bg="#1A6634" if "21" in h else "#7B3F00" if "22" in h else "#2C3E50", sz=9)
    for i in range(min(20, len(wordfreq["tfidf_distinctive_participants_fl21"]))):
        alt = (i%2==0)
        dat(ws4.cell(row=tfidf_row+2+i,column=1), i+1, alt=alt, bold=True)
        w21, _, d21 = wordfreq["tfidf_distinctive_participants_fl21"][i]
        w22, _, d22 = wordfreq["tfidf_distinctive_participants_fl22"][i]
        dat(ws4.cell(row=tfidf_row+2+i,column=2), w21, center=False, alt=alt)
        dat(ws4.cell(row=tfidf_row+2+i,column=3), d21, alt=alt, fmt="0.0000")
        dat(ws4.cell(row=tfidf_row+2+i,column=4), w22, center=False, alt=alt)
        dat(ws4.cell(row=tfidf_row+2+i,column=5), d22, alt=alt, fmt="0.0000")

    insert_png(ws4, OUTPUT_DIR/"fig_topwords_participants.png", "I2", 22, 9)
    insert_png(ws4, OUTPUT_DIR/"fig_tfidf_participants.png",   "I22",22, 9)
    insert_png(ws4, OUTPUT_DIR/"fig_wc_part_fl21.png",         "A50",16, 7)
    insert_png(ws4, OUTPUT_DIR/"fig_wc_part_fl22.png",         "J50",16, 7)
    for j, w in enumerate([6,22,10,22,10,10], 1):
        ws4.column_dimensions[get_column_letter(j)].width = w
    print("Feuille 4 — Mots participants OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 5 — Top mots chatbot
    # ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("5_Mots_Chatbot")
    title_row(ws5, "Fréquences de mots & TF-IDF — Chatbot FL_21 vs FL_22", "#7B3F00", 8)
    max_rb = max(len(wordfreq["top_words_chatbot_fl21"]), len(wordfreq["top_words_chatbot_fl22"]))
    for j, h in enumerate(["Rang","Mot Chatbot FL_21","Freq","Mot Chatbot FL_22","Freq","En commun"], 1):
        hdr(ws5.cell(row=2,column=j,value=h),
            bg="#1A6634" if "21" in h else "#7B3F00" if "22" in h else "#2C3E50", sz=9)
    m22_bot = {w for w,_ in wordfreq["top_words_chatbot_fl22"]}
    for i in range(max_rb):
        alt = (i%2==0)
        dat(ws5.cell(row=3+i,column=1), i+1, alt=alt, bold=True)
        if i < len(wordfreq["top_words_chatbot_fl21"]):
            w21, c21 = wordfreq["top_words_chatbot_fl21"][i]
            dat(ws5.cell(row=3+i,column=2), w21, center=False, alt=alt)
            dat(ws5.cell(row=3+i,column=3), c21, alt=alt)
            dat(ws5.cell(row=3+i,column=6), "✓" if w21 in m22_bot else "", alt=alt)
        if i < len(wordfreq["top_words_chatbot_fl22"]):
            w22, c22 = wordfreq["top_words_chatbot_fl22"][i]
            dat(ws5.cell(row=3+i,column=4), w22, center=False, alt=alt)
            dat(ws5.cell(row=3+i,column=5), c22, alt=alt)
    tfidf_row5 = max_rb + 5
    title_row(ws5, "Mots distinctifs TF-IDF — Chatbot", "#C0392B", 8, row=tfidf_row5)
    for j, h in enumerate(["Rang","Distinctif Chatbot FL_21","Δ TF-IDF","Distinctif Chatbot FL_22","Δ TF-IDF"], 1):
        hdr(ws5.cell(row=tfidf_row5+1,column=j,value=h),
            bg="#1A6634" if "21" in h else "#7B3F00" if "22" in h else "#2C3E50", sz=9)
    for i in range(min(20, len(wordfreq["tfidf_distinctive_chatbot_fl21"]))):
        alt = (i%2==0)
        dat(ws5.cell(row=tfidf_row5+2+i,column=1), i+1, alt=alt, bold=True)
        w21, _, d21 = wordfreq["tfidf_distinctive_chatbot_fl21"][i]
        w22, _, d22 = wordfreq["tfidf_distinctive_chatbot_fl22"][i]
        dat(ws5.cell(row=tfidf_row5+2+i,column=2), w21, center=False, alt=alt)
        dat(ws5.cell(row=tfidf_row5+2+i,column=3), d21, alt=alt, fmt="0.0000")
        dat(ws5.cell(row=tfidf_row5+2+i,column=4), w22, center=False, alt=alt)
        dat(ws5.cell(row=tfidf_row5+2+i,column=5), d22, alt=alt, fmt="0.0000")
    insert_png(ws5, OUTPUT_DIR/"fig_topwords_chatbot.png",   "I2",  22, 9)
    insert_png(ws5, OUTPUT_DIR/"fig_tfidf_chatbot.png",      "I22", 22, 9)
    insert_png(ws5, OUTPUT_DIR/"fig_wc_bot_fl21.png",        "A50", 16, 7)
    insert_png(ws5, OUTPUT_DIR/"fig_wc_bot_fl22.png",        "J50", 16, 7)
    for j, w in enumerate([6,22,10,22,10,10], 1):
        ws5.column_dimensions[get_column_letter(j)].width = w
    print("Feuille 5 — Mots chatbot OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 6 — Analyse IA : comparaison
    # ──────────────────────────────────────────────────
    ws6 = wb.create_sheet("6_Analyse_IA_Comparaison")
    if df_ai.empty:
        title_row(ws6, "Résultats IA non disponibles — lancez 03_api_analysis.py", "#888888", 4)
    else:
        title_row(ws6, "Analyse IA — Qualité & Actionabilité des réponses FL_21 vs FL_22 (Claude)", "#4A235A", 9)
        write_ttest_table(ws6, ai_content_ttests, start_row=2)
        insert_png(ws6, OUTPUT_DIR/"fig_ai_content.png", "K2", 22, 9)
    print("Feuille 6 — Analyse IA comparaison OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 7 — Détail IA par participant
    # ──────────────────────────────────────────────────
    ws7 = wb.create_sheet("7_Détail_IA_Participants")
    title_row(ws7, "Classification IA détaillée — une ligne par participant", "#4A235A", 20)
    if not df_ai.empty:
        COLS7 = ["respondent_id","version","n_turns",
                 "qualite_score","qualite_label","action_score","action_pb_concret","action_conseil","action_usecase",
                 "engagement","elaboration","coherence","expertise",
                 "has_opinion","has_frustration","has_suggestion","has_experience","has_feature_req",
                 "sentiment","langue","abandon","themes","resume","verbatim"]
        COLS7 = [c for c in COLS7 if c in df_ai.columns]
        for j, col in enumerate(COLS7, 1):
            hdr(ws7.cell(row=2,column=j,value=col), bg="#4A235A", sz=8)
        ws7.row_dimensions[2].height = 28
        SCORE_COLORS = {1:"FF4444",2:"FF8C00",3:"FFD700",4:"4CAF50",5:"1A6634"}
        SCORE_FC     = {1:"FFFFFF",2:"FFFFFF",3:"000000",4:"FFFFFF",5:"FFFFFF"}
        for i, (_, row_ai) in enumerate(df_ai.sort_values(["version","qualite_score"],ascending=[True,False]).iterrows()):
            r = i+3; alt=(i%2==0)
            for j, col in enumerate(COLS7, 1):
                v   = row_ai.get(col)
                c   = ws7.cell(row=r, column=j)
                if col in ["qualite_score","action_score"] and pd.notna(v):
                    vi = int(v)
                    c.value = vi; c.border = THIN
                    c.alignment = Alignment(horizontal="center",vertical="center")
                    c.fill = PatternFill("solid", start_color=SCORE_COLORS.get(vi,"FFFFFF"))
                    c.font = Font(name="Arial",size=9,bold=True,color=SCORE_FC.get(vi,"000000"))
                elif col == "version":
                    c.value = v; c.border = THIN
                    c.alignment = Alignment(horizontal="center",vertical="center")
                    c.fill = PatternFill("solid", start_color="1A6634" if v=="FL_21" else "7B3F00")
                    c.font = Font(name="Arial",size=9,bold=True,color="FFFFFF")
                else:
                    dat(c, v, center=(col not in ["respondent_id","resume","verbatim","themes"]), alt=alt)
            ws7.row_dimensions[r].height = 28
        for j, col in enumerate(COLS7, 1):
            w = 45 if col in ["resume","verbatim","themes"] else 22 if col in ["respondent_id"] else 14
            ws7.column_dimensions[get_column_letter(j)].width = w
        ws7.freeze_panes = "A3"
    print("Feuille 7 — Détail IA participants OK")

    # ──────────────────────────────────────────────────
    # FEUILLE 8 — Style chatbot (analyse IA)
    # ──────────────────────────────────────────────────
    ws8 = wb.create_sheet("8_Style_Chatbot_IA")
    title_row(ws8, "Analyse IA du style du chatbot — FL_21 (Friendly) vs FL_22 (Pro)", "#1A5276", 12)
    if not df_ai.empty:
        bot_ttests = [ttest_binary_ai(df_ai, col, lbl) for col, lbl in [
            ("bot_chaleur",      "Score chaleur perçue (1-5)"),
            ("bot_formalisme",   "Score formalisme (1-5)"),
            ("bot_empathie",     "Score empathie (1-5)"),
            ("bot_pct_ouvertes", "% questions ouvertes"),
            ("bot_relances",     "Relances personnalisées"),
            ("bot_encouragement","Formules d'encouragement"),
        ]]
        write_ttest_table(ws8, bot_ttests, start_row=2)

        # Marqueurs de ton : tableau
        bot_row = len(bot_ttests) + 5
        title_row(ws8, "Marqueurs de ton identifiés par l'IA (expressions caractéristiques)", "#1A5276", 6, row=bot_row)
        bot_row += 1
        for j, h in enumerate(["Version","Score chaleur","Score formalisme","Ton label","Marqueurs Friendly","Marqueurs Pro"], 1):
            hdr(ws8.cell(row=bot_row, column=j, value=h), bg="#2C3E50", sz=9)
        bot_row += 1
        bot_cols = ["version","bot_chaleur","bot_formalisme","bot_ton_label","bot_friendly_markers","bot_pro_markers"]
        bot_cols = [c for c in bot_cols if c in df_ai.columns]
        for i, (_, row_b) in enumerate(df_ai.sort_values("version").iterrows()):
            r = bot_row + i; alt = (i%2==0)
            for j, col in enumerate(bot_cols, 1):
                c = ws8.cell(row=r, column=j)
                if col == "version":
                    c.value = row_b[col]; c.border = THIN
                    c.alignment = Alignment(horizontal="center",vertical="center")
                    c.fill = PatternFill("solid", start_color="1A6634" if row_b[col]=="FL_21" else "7B3F00")
                    c.font = Font(name="Arial",size=9,bold=True,color="FFFFFF")
                else:
                    dat(c, row_b.get(col), center=(col not in ["bot_friendly_markers","bot_pro_markers"]), alt=alt)
            ws8.row_dimensions[r].height = 25
        for j, w in enumerate([10,12,14,18,40,40], 1):
            ws8.column_dimensions[get_column_letter(j)].width = w
    print("Feuille 8 — Style chatbot IA OK")

    # ── Sauvegarde ────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\nFichier Excel final : {OUTPUT_FILE}")
    return str(OUTPUT_FILE)

if __name__ == "__main__":
    import sys, importlib
    run()
PYEOF
Sortie

exit code 0
